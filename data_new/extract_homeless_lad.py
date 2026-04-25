from pathlib import Path
import argparse
import csv
import re
from openpyxl import load_workbook

OUTPUT_COLUMNS = [
    "LAD_code",
    "LA_name",
    "Year",
    "Quarter",
    "Homeless_relief",
    "Total_assessments",
    "Homeless_per_1000",
]

LAD_RE = re.compile(r"^E0[6-9]\d{6}$")
YEAR_RE = re.compile(r"\b(20\d{2})\b")
QUARTER_PATTERNS = [
    (re.compile(r"january\s*(?:-|to|–|—)\s*march|jan\w*\s*(?:-|to|–|—)\s*mar", re.I), "Q1"),
    (re.compile(r"april\s*(?:-|to|–|—)\s*june|apr\w*\s*(?:-|to|–|—)\s*jun", re.I), "Q2"),
    (re.compile(r"july\s*(?:-|to|–|—)\s*september|jul\w*\s*(?:-|to|–|—)\s*sep", re.I), "Q3"),
    (re.compile(r"october\s*(?:-|to|–|—)\s*december|oct\w*\s*(?:-|to|–|—)\s*dec", re.I), "Q4"),
]
SHEET_QUARTERS = {
    "mar": "Q1", "march": "Q1",
    "jun": "Q2", "june": "Q2",
    "sep": "Q3", "sept": "Q3", "september": "Q3",
    "dec": "Q4", "december": "Q4",
}


def norm_text(x):
    if x is None:
        return ""
    s = str(x).replace("\n", " ").replace("\r", " ")
    s = s.replace("–", "-").replace("—", "-").replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def clean_value(v):
    if v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s in {"", "-", "--", "..", ".", "*", "x", "X"}:
            return ""
        s_num = s.replace(",", "")
        if re.fullmatch(r"-?\d+(\.\d+)?", s_num):
            f = float(s_num)
            return int(f) if f.is_integer() else f
        return s
    return v


def parse_period(*texts):
    text = " ".join(str(t) for t in texts if t)
    year_match = YEAR_RE.search(text)
    year = int(year_match.group(1)) if year_match else None
    quarter = None
    normalized = text.replace(" - ", " to ").replace("-", " to ")
    for pat, q in QUARTER_PATTERNS:
        if pat.search(normalized):
            quarter = q
            break
    if quarter is None:
        m = re.search(r"(mar|march|jun|june|sep|sept|september|dec|december)[_\s-]*(20\d{2})", text, re.I)
        if m:
            quarter = SHEET_QUARTERS[m.group(1).lower()]
            year = year or int(m.group(2))
    if year is None or quarter is None:
        return None, None
    return year, quarter


def first_rows(ws, max_row=10):
    return list(ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row), max_col=ws.max_column, values_only=True))


def combined_headers(rows, start_row=1, end_row=7):
    headers = {}
    max_col = max((len(r) for r in rows), default=0)
    for c in range(max_col):
        parts = []
        for r in range(start_row - 1, min(end_row, len(rows))):
            v = rows[r][c] if c < len(rows[r]) else None
            if v not in (None, ""):
                parts.append(str(v))
        headers[c + 1] = norm_text(" ".join(parts))
    return headers


def find_col(headers, predicates, min_col=1, max_col=None):
    candidates = []
    for c, h in headers.items():
        if c < min_col:
            continue
        if max_col is not None and c > max_col:
            continue
        if all(p in h for p in predicates):
            candidates.append(c)
    return min(candidates) if candidates else None


def make_record(code, name, year, quarter, homeless_relief="", total_assessments="", homeless_per_1000="", priority=0, source=""):
    return {
        "LAD_code": code,
        "LA_name": name,
        "Year": year,
        "Quarter": quarter,
        "Homeless_relief": clean_value(homeless_relief),
        "Total_assessments": clean_value(total_assessments),
        "Homeless_per_1000": clean_value(homeless_per_1000),
        "_priority": priority,
        "_source": source,
    }


def extract_old_p1e_file(path):
    records = []
    year, quarter = parse_period(path.stem)
    if not year or not quarter:
        return records
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Section 1" not in wb.sheetnames:
        wb.close()
        return records
    ws = wb["Section 1"]
    header_row_idx = None
    code_col = name_col = total_col = None
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        vals = [norm_text(v) for v in row]
        if "ons code" in vals and "local authority" in vals:
            header_row_idx = ridx
            code_col = vals.index("ons code") + 1
            name_col = vals.index("local authority") + 1
            for i, v in enumerate(row, start=1):
                if norm_text(v) == "e16w":
                    total_col = i
                    break
            break
    if not header_row_idx or not total_col:
        wb.close()
        return records
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        code = clean_value(row[code_col - 1] if len(row) >= code_col else "")
        if not isinstance(code, str) or not LAD_RE.match(code):
            continue
        name = clean_value(row[name_col - 1] if len(row) >= name_col else "")
        total = row[total_col - 1] if len(row) >= total_col else ""
        records.append(make_record(code, name, year, quarter, "", total, "", priority=10, source=path.name))
    wb.close()
    return records


def extract_summary_14_18(path):
    records = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        if not ws.title.lower().endswith("_qtr"):
            continue
        rows = first_rows(ws, 8)
        title = rows[0][0] if rows and rows[0] else ""
        year, quarter = parse_period(title, ws.title)
        if not year or not quarter:
            continue
        headers = combined_headers(rows, 1, 6)
        total_col = find_col(headers, ["total decisions"])
        per_cols = [c for c, h in headers.items() if "number per" in h and ("1000" in h or "1,000" in h)]
        per_col = min([c for c in per_cols if total_col is None or c < total_col], default=None)
        if total_col is None:
            continue
        for row in ws.iter_rows(min_row=7, values_only=True):
            code = clean_value(row[0] if len(row) > 0 else "")
            if not isinstance(code, str) or not LAD_RE.match(code):
                continue
            name = clean_value(row[1] if len(row) > 1 else "")
            total = row[total_col - 1] if len(row) >= total_col else ""
            per1000 = row[per_col - 1] if per_col and len(row) >= per_col else ""
            records.append(make_record(code, name, year, quarter, "", total, per1000, priority=20, source=f"{path.name}:{ws.title}"))
    wb.close()
    return records


def extract_hclic_18_25(path):
    records = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = first_rows(ws, 8)
        title = rows[0][0] if rows and rows[0] else ""
        if "table a1" not in norm_text(title):
            continue
        year, quarter = parse_period(title, ws.title)
        if not year or not quarter:
            continue
        headers = combined_headers(rows, 1, 7)
        total_col = find_col(headers, ["total initial assessments"]) or find_col(headers, ["total number of households assessed"])
        relief_col = None
        relief_candidates = []
        for c, h in headers.items():
            if "relief duty owed" in h and "per" not in h and "section 21" not in h:
                relief_candidates.append(c)
        if relief_candidates:
            relief_col = min(relief_candidates)
        per_col = None
        per_candidates = []
        for c, h in headers.items():
            if "assessed as homeless" in h and "per" in h and "000" in h:
                per_candidates.append(c)
        if per_candidates:
            per_col = min(per_candidates)
        if total_col is None and relief_col is None:
            continue
        for row in ws.iter_rows(min_row=6, values_only=True):
            code = clean_value(row[0] if len(row) > 0 else "")
            if not isinstance(code, str) or not LAD_RE.match(code):
                continue
            name = clean_value(row[1] if len(row) > 1 else "")
            total = row[total_col - 1] if total_col and len(row) >= total_col else ""
            relief = row[relief_col - 1] if relief_col and len(row) >= relief_col else ""
            per1000 = row[per_col - 1] if per_col and len(row) >= per_col else ""
            records.append(make_record(code, name, year, quarter, relief, total, per1000, priority=30, source=f"{path.name}:{ws.title}"))
    wb.close()
    return records


def dedupe(records):
    by_key = {}
    for rec in records:
        key = (rec["LAD_code"], int(rec["Year"]), rec["Quarter"])
        if key not in by_key:
            by_key[key] = rec.copy()
            continue
        old = by_key[key]
        if rec["_priority"] >= old["_priority"]:
            merged = rec.copy()
            for col in OUTPUT_COLUMNS:
                if merged.get(col, "") in (None, "") and old.get(col, "") not in (None, ""):
                    merged[col] = old[col]
            by_key[key] = merged
        else:
            for col in OUTPUT_COLUMNS:
                if old.get(col, "") in (None, "") and rec.get(col, "") not in (None, ""):
                    old[col] = rec[col]
    return list(by_key.values())


def collect_records(data_dir):
    data_dir = Path(data_dir)
    records = []
    for path in sorted(data_dir.rglob("*.xlsx")):
        lower = path.name.lower()
        if lower.startswith("~$"):
            continue
        if "homelessness 14-18" in lower or "homelessness 14" in lower:
            records.extend(extract_summary_14_18(path))
        elif "homelessness 18-25" in lower or "homelessness 18" in lower:
            records.extend(extract_hclic_18_25(path))
        else:
            records.extend(extract_old_p1e_file(path))
    return dedupe(records)


def sort_key(rec):
    q_order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    return (int(rec["Year"]), q_order.get(rec["Quarter"], 99), rec["LAD_code"])


def write_csv(records, output_path):
    output_path = Path(output_path)
    records = sorted(records, key=sort_key)
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for rec in records:
            writer.writerow({col: rec.get(col, "") for col in OUTPUT_COLUMNS})


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-dir", default="data", help="folder containing the xlsx files")
    parser.add_argument("--output", default="homeless_lad_2009_2025.csv")
    args = parser.parse_args()
    records = collect_records(args.data_dir)
    write_csv(records, args.output)
    print(f"wrote {len(records):,} rows to {args.output}")


if __name__ == "__main__":
    main()
